from flask import Flask, request, jsonify, render_template, redirect, url_for, session
import sqlite3
import openpyxl
import logging
import os
import sys
import tempfile
import pandas as pd
import shutil

# 导入蓝图
from routes.devices import devices_bp
from routes.physical import physical_bp
from routes.department import department_bp

# 获取持久化目录的公共函数
def get_persist_dir():
    """获取持久化目录 - 统一使用项目目录下的data文件夹"""
    # 获取当前文件所在目录（EMS1.4/）
    current_dir = os.path.dirname(os.path.abspath(__file__))
    # 使用项目目录下的data文件夹
    persist_dir = os.path.join(current_dir, 'data')

    # 确保持久化目录存在
    os.makedirs(persist_dir, exist_ok=True)
    return persist_dir

# 获取资源目录（打包后的资源目录）
def get_resource_dir():
    if getattr(sys, 'frozen', False):
        # 打包后的可执行文件
        return sys._MEIPASS
    else:
        # 未打包的脚本
        return os.path.dirname(os.path.abspath(__file__))

# 从打包的资源中提取数据库文件（仅在打包环境下执行）
def extract_database_from_bundle():
    # 如果不是打包环境，直接返回False
    if not getattr(sys, 'frozen', False):
        return False
    
    persist_dir = get_persist_dir()
    db_path = os.path.join(persist_dir, 'device_management.db')
    init_flag_path = os.path.join(persist_dir, 'init_done.flag')
    
    # 如果数据库文件不存在，从打包的资源中提取
    if not os.path.exists(db_path) or not os.path.exists(init_flag_path):
        print("首次运行，从打包资源中提取数据库文件...")
        resource_dir = get_resource_dir()
        bundled_db_path = os.path.join(resource_dir, 'data', 'device_management.db')
        
        if os.path.exists(bundled_db_path):
            print(f"从 {bundled_db_path} 复制数据库文件到 {db_path}")
            shutil.copy2(bundled_db_path, db_path)
            print("数据库文件提取成功")
            
            # 创建标志文件
            with open(init_flag_path, 'w') as f:
                f.write('初始化完成')
            print(f"创建初始化标志文件成功: {init_flag_path}")
        else:
            print(f"警告: 打包资源中未找到数据库文件: {bundled_db_path}")
            return False
    
    return True

# 导入Dash相关库
from dash import Dash, dcc, html, Input, Output

# 数据库初始化和重置逻辑
try:
    # 获取当前目录
    current_dir = os.path.dirname(os.path.abspath(__file__))
    init_db_path = os.path.join(current_dir, 'init_db.py')
    
    # 确定持久化目录
    persist_dir = get_persist_dir()
    
    # 标志文件路径
    init_flag_path = os.path.join(persist_dir, 'init_done.flag')
    
    # 数据库文件路径（存储在持久化目录）
    db_path = os.path.join(persist_dir, 'device_management.db')
    
    # 确保 data 目录存在（保持兼容性）
    data_dir = os.path.join(current_dir, 'data')
    os.makedirs(data_dir, exist_ok=True)
    
    # 检查是否是第一次启动
    if not os.path.exists(init_flag_path):
        print("首次启动，尝试从打包资源中提取数据库...")
        # 先尝试从打包资源中提取数据库
        if not extract_database_from_bundle():
            print("从打包资源提取数据库失败，执行数据库初始化脚本...")
            # 检查 init_db.py 文件是否存在
            if os.path.exists(init_db_path):
                print(f"执行数据库初始化脚本: {init_db_path}")
                # 执行初始化脚本
                with open(init_db_path, 'r', encoding='utf-8') as f:
                    init_db_code = f.read()
                exec(init_db_code)
                print("数据库初始化脚本执行完成")
                # 创建标志文件，标记初始化完成
                with open(init_flag_path, 'w') as f:
                    f.write('初始化完成')
                print(f"创建初始化标志文件成功: {init_flag_path}")
            else:
                print(f"警告: 未找到数据库初始化脚本: {init_db_path}")
    else:
        # 检查是否是Flask的自动重新加载过程
        if os.environ.get('WERKZEUG_RUN_MAIN') != 'true':
            # 非首次启动，使用现有数据库
            print("非首次启动，检测到现有数据库")
            print(f"数据库文件位置: {db_path}")
            
            # 自动重置数据库，因为我们修改了数据库结构
            print("自动重置数据库以应用新的数据库结构...")
            # 执行初始化脚本（会重置数据库）
            if os.path.exists(init_db_path):
                print(f"执行数据库初始化脚本: {init_db_path}")
                with open(init_db_path, 'r', encoding='utf-8') as f:
                    init_db_code = f.read()
                exec(init_db_code)
                print("数据库重置完成")
            else:
                print(f"警告: 未找到数据库初始化脚本: {init_db_path}")
        else:
            # Flask重新加载过程，跳过用户交互
            print("Flask重新加载过程，跳过数据库重置询问")
            print("使用现有数据库，保留之前的设备数据")
except Exception as e:
    print(f"数据库初始化检查时出错: {str(e)}")
    # 继续执行，不中断应用启动

# 配置日志
# 确定日志文件路径
# 获取可执行文件所在的目录
if getattr(sys, 'frozen', False):
    # 打包后的可执行文件
    executable_dir = os.path.dirname(sys.executable)
else:
    # 未打包的脚本
    executable_dir = os.path.dirname(os.path.abspath(__file__))

# 日志文件路径 - 直接放在可执行文件旁边
log_file = os.path.join(executable_dir, 'ems.log')

# 日志目录就是可执行文件所在目录，不需要额外创建

# 配置日志，同时输出到文件和终端
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file),
        logging.StreamHandler()
    ]
)
print(f"日志文件位置: {log_file}")

app = Flask(__name__)
app.config['JSON_AS_ASCII'] = False
app.config['SECRET_KEY'] = 'your-secret-key-here'  # 用于会话管理

# 添加CORS支持
@app.after_request
def add_cors_headers(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
    response.headers['Access-Control-Allow-Credentials'] = 'true'
    return response

# 处理OPTIONS请求
@app.route('/api/<path:path>', methods=['OPTIONS'])
def handle_options(path):
    return '', 204

# 注册蓝图
app.register_blueprint(devices_bp)
app.register_blueprint(physical_bp)
app.register_blueprint(department_bp)

# 创建Dash应用，绑定到Flask服务器
dash_app = Dash(__name__, server=app, url_base_pathname='/dashboard/')

# 数据库连接函数
def get_db_connection():
    import os
    # 使用持久化目录中的数据库文件
    persist_dir = get_persist_dir()
    
    # 数据库文件路径
    db_path = os.path.join(persist_dir, 'device_management.db')
    print(f"连接数据库: {db_path}")
    
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn

# 登录页面
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        # 从数据库获取用户信息
        conn = get_db_connection()
        user = conn.execute('SELECT users.*, roles.name as role_name FROM users JOIN roles ON users.role_id = roles.id WHERE users.username = ?', (username,)).fetchone()
        conn.close()
        
        # 验证用户名和密码
        if user and user['password'] == password:
            # 登录成功，设置会话
            session['logged_in'] = True
            session['username'] = username
            session['role'] = user['role_name']
            session['user_id'] = user['id']
            session['department'] = user['department']
            return redirect(url_for('index'))
        else:
            # 登录失败，重定向到登录页面
            return render_template('login.html', error='账号或密码错误')
    
    # GET请求，显示登录页面
    return render_template('login.html')

# API登录端点（返回JSON格式）
@app.route('/api/login', methods=['POST'])
def api_login():
    username = request.form.get('username')
    password = request.form.get('password')
    
    # 从数据库获取用户信息
    conn = get_db_connection()
    user = conn.execute('SELECT users.*, roles.name as role_name FROM users JOIN roles ON users.role_id = roles.id WHERE users.username = ?', (username,)).fetchone()
    conn.close()
    
    # 验证用户名和密码
    if user and user['password'] == password:
        # 登录成功，设置会话
            session['logged_in'] = True
            session['username'] = username
            session['role'] = user['role_name']
            session['user_id'] = user['id']
            session['department'] = user['department']
            
            # 返回JSON格式的成功信息
            return jsonify({
                'success': True,
                'message': '登录成功',
                'user': {
                    'username': user['username'],
                    'role': user['role_name']
                }
            })
    else:
        # 返回JSON格式的错误信息
        return jsonify({
            'success': False,
            'message': '账号或密码错误'
        }), 401

# 登出路由
@app.route('/logout')
def logout():
    # 清除会话
    session.clear()
    return redirect(url_for('login'))

# API登出端点（返回JSON格式）
@app.route('/api/logout', methods=['POST'])
def api_logout():
    # 清除会话
    session.clear()
    
    # 返回JSON格式的成功信息
    return jsonify({
        'success': True,
        'message': '登出成功'
    })

# 登录检查装饰器
from functools import wraps
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            # 检查是否是API请求
            if request.path.startswith('/api/'):
                # API请求返回JSON格式的错误信息
                return jsonify({'error': '未登录，请先登录'}), 401
            else:
                # 非API请求重定向到登录页面
                return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# 管理员权限检查装饰器
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            # 检查是否是API请求
            if request.path.startswith('/api/'):
                # API请求返回JSON格式的错误信息
                return jsonify({'error': '未登录，请先登录'}), 401
            else:
                # 非API请求重定向到登录页面
                return redirect(url_for('login'))
        if session.get('role') != 'admin':
            # 检查是否是API请求
            if request.path.startswith('/api/'):
                # API请求返回JSON格式的错误信息
                return jsonify({'error': '权限不足，需要管理员权限'}), 403
            else:
                # 非API请求返回错误页面或重定向
                return render_template('error.html', message='权限不足，需要管理员权限'), 403
        return f(*args, **kwargs)
    return decorated_function

# 系统说明页面
@app.route('/help')
@login_required
def system_info():
    return render_template('help.html')

# 账号管理页面
@app.route('/user_management')
@login_required
@admin_required
def user_management():
    return render_template('user_management.html', active_page='user_management')



# 根路由，返回前端界面
@app.route('/')
@login_required
def index():
    return render_template('index.html', active_page='vm')

# 物理机管理页面
@app.route('/physical')
@login_required
def physical():
    return render_template('physical.html', active_page='physical')

@app.route('/department')
@login_required
def department():
    # 仅管理员可见
    if session.get('role') != 'admin':
        flash('无权限访问此页面', 'danger')
        return redirect(url_for('index'))
    return render_template('department.html', active_page='department')

# 监控检查页面
@app.route('/check')
@login_required
def check():
    return render_template('check.html', active_page='check')

















# 监控检查页面
@app.route('/check')
@login_required
def check_devices_page():
    return render_template('check.html')

# 数据统计页面
@app.route('/stats')
@login_required
@admin_required
def stats():
    conn = get_db_connection()
    
    # 计算虚拟机数量
    vm_count = conn.execute('SELECT COUNT(*) FROM devices').fetchone()[0]
    
    # 计算物理机数量
    physical_count = conn.execute('SELECT COUNT(*) FROM physical_devices').fetchone()[0]
    
    # 计算产品数量（去重，包括虚拟机和物理机）
    vm_product_count = conn.execute('SELECT COUNT(DISTINCT product) FROM devices WHERE product IS NOT NULL AND product != ""').fetchone()[0]
    physical_product_count = conn.execute('SELECT COUNT(DISTINCT product) FROM physical_devices WHERE product IS NOT NULL AND product != ""').fetchone()[0]
    # 合并产品列表并去重
    vm_products = conn.execute('SELECT DISTINCT product FROM devices WHERE product IS NOT NULL AND product != ""').fetchall()
    physical_products = conn.execute('SELECT DISTINCT product FROM physical_devices WHERE product IS NOT NULL AND product != ""').fetchall()
    all_products = set()
    for row in vm_products:
        if row['product']:
            all_products.add(row['product'])
    for row in physical_products:
        if row['product']:
            all_products.add(row['product'])
    product_count = len(all_products)
    
    # 计算部门数量（去重）
    department_count = conn.execute('SELECT COUNT(DISTINCT department) FROM devices WHERE department IS NOT NULL AND department != ""').fetchone()[0]
    physical_department_count = conn.execute('SELECT COUNT(DISTINCT department) FROM physical_devices WHERE department IS NOT NULL AND department != ""').fetchone()[0]
    department_count = max(department_count, physical_department_count)
    
    # 计算组别数量（去重，考虑部门和组别的组合）
    # 从devices表获取部门和组别的唯一组合
    vm_groups = conn.execute('SELECT DISTINCT department, device_group FROM devices WHERE department IS NOT NULL AND department != "" AND device_group IS NOT NULL AND device_group != ""').fetchall()
    # 合并并去重
    all_groups = set()
    for row in vm_groups:
        all_groups.add((row['department'], row['device_group']))
    group_count = len(all_groups)
    
    # 计算每个产品使用的机器个数（包括虚拟机和物理机）
    vm_product_data = conn.execute('''
        SELECT product, COUNT(*) as count 
        FROM devices 
        WHERE product IS NOT NULL AND product != ""
        GROUP BY product
    ''').fetchall()
    
    physical_product_data = conn.execute('''
        SELECT product, COUNT(*) as count 
        FROM physical_devices 
        WHERE product IS NOT NULL AND product != ""
        GROUP BY product
    ''').fetchall()
    
    # 合并产品数据
    product_dict = {}
    for row in vm_product_data:
        product_dict[row['product']] = product_dict.get(row['product'], 0) + row['count']
    for row in physical_product_data:
        product_dict[row['product']] = product_dict.get(row['product'], 0) + row['count']
    
    # 转换为列表并排序
    product_data = [{'product': product, 'count': count} for product, count in product_dict.items()]
    product_data.sort(key=lambda x: x['count'], reverse=True)
    
    # 计算每个部门使用的机器个数（包括虚拟机和物理机）
    # 先获取虚拟机部门数据
    vm_department_data = conn.execute('''
        SELECT department, COUNT(*) as count 
        FROM devices 
        WHERE department IS NOT NULL AND department != ""
        GROUP BY department
    ''').fetchall()
    
    # 再获取物理机部门数据
    physical_department_data = conn.execute('''
        SELECT department, COUNT(*) as count 
        FROM physical_devices 
        WHERE department IS NOT NULL AND department != ""
        GROUP BY department
    ''').fetchall()
    
    # 合并部门数据
    department_dict = {}
    for row in vm_department_data:
        department_dict[row['department']] = department_dict.get(row['department'], 0) + row['count']
    for row in physical_department_data:
        department_dict[row['department']] = department_dict.get(row['department'], 0) + row['count']
    
    # 转换为列表并排序
    department_data = [{'department': dept, 'count': count} for dept, count in department_dict.items()]
    department_data.sort(key=lambda x: x['count'], reverse=True)
    
    # 准备虚拟机按部门分布数据（用于饼图）
    vm_by_department = [{'department': row['department'], 'count': row['count']} for row in vm_department_data]
    vm_by_department.sort(key=lambda x: x['count'], reverse=True)
    
    # 准备物理机按部门分布数据（用于饼图）
    physical_by_department = [{'department': row['department'], 'count': row['count']} for row in physical_department_data]
    physical_by_department.sort(key=lambda x: x['count'], reverse=True)
    
    # 计算每个部门的产品个数
    department_product_data = conn.execute('''
        SELECT department, COUNT(DISTINCT product) as product_count 
        FROM devices 
        WHERE department IS NOT NULL AND department != "" AND product IS NOT NULL AND product != ""
        GROUP BY department 
        ORDER BY product_count DESC
    ''').fetchall()
    department_product_data = [dict(row) for row in department_product_data]
    
    # 计算设备类型分布
    device_type_data = [
        {'type': '虚拟机', 'count': vm_count},
        {'type': '物理机', 'count': physical_count}
    ]
    
    # 计算部门-产品资源矩阵（包括虚拟机和物理机）
    vm_matrix_data = conn.execute('''
        SELECT department, product, COUNT(*) as count 
        FROM devices 
        WHERE department IS NOT NULL AND department != "" AND product IS NOT NULL AND product != ""
        GROUP BY department, product
    ''').fetchall()
    
    physical_matrix_data = conn.execute('''
        SELECT department, product, COUNT(*) as count 
        FROM physical_devices 
        WHERE department IS NOT NULL AND department != "" AND product IS NOT NULL AND product != ""
        GROUP BY department, product
    ''').fetchall()
    
    # 合并部门-产品矩阵数据
    matrix_dict = {}
    for row in vm_matrix_data:
        key = (row['department'], row['product'])
        matrix_dict[key] = matrix_dict.get(key, 0) + row['count']
    for row in physical_matrix_data:
        key = (row['department'], row['product'])
        matrix_dict[key] = matrix_dict.get(key, 0) + row['count']
    
    # 转换为列表
    department_product_matrix = [{'department': key[0], 'product': key[1], 'count': count} for key, count in matrix_dict.items()]
    
    # 计算资源使用趋势（模拟数据，实际应根据创建时间统计）
    # 月度数据
    resource_trend_data = [
        {'month': '1月', 'vm_count': vm_count // 12 * 8, 'physical_count': physical_count // 12 * 8},
        {'month': '2月', 'vm_count': vm_count // 12 * 8, 'physical_count': physical_count // 12 * 8},
        {'month': '3月', 'vm_count': vm_count // 12 * 9, 'physical_count': physical_count // 12 * 9},
        {'month': '4月', 'vm_count': vm_count // 12 * 9, 'physical_count': physical_count // 12 * 9},
        {'month': '5月', 'vm_count': vm_count // 12 * 10, 'physical_count': physical_count // 12 * 10},
        {'month': '6月', 'vm_count': vm_count // 12 * 10, 'physical_count': physical_count // 12 * 10},
        {'month': '7月', 'vm_count': vm_count // 12 * 11, 'physical_count': physical_count // 12 * 11},
        {'month': '8月', 'vm_count': vm_count // 12 * 11, 'physical_count': physical_count // 12 * 11},
        {'month': '9月', 'vm_count': vm_count // 12 * 12, 'physical_count': physical_count // 12 * 12},
        {'month': '10月', 'vm_count': vm_count, 'physical_count': physical_count},
        {'month': '11月', 'vm_count': vm_count, 'physical_count': physical_count},
        {'month': '12月', 'vm_count': vm_count, 'physical_count': physical_count}
    ]
    
    # 季度数据（显示四个季度，数据保持一致）
    resource_quarterly_data = [
        {'quarter': 'Q1', 'vm_count': vm_count, 'physical_count': physical_count},
        {'quarter': 'Q2', 'vm_count': vm_count, 'physical_count': physical_count},
        {'quarter': 'Q3', 'vm_count': vm_count, 'physical_count': physical_count},
        {'quarter': 'Q4', 'vm_count': vm_count, 'physical_count': physical_count}
    ]
    
    # 年份数据（使用当前年份）
    import datetime
    current_year = str(datetime.datetime.now().year)
    resource_yearly_data = [
        {'year': current_year, 'vm_count': vm_count, 'physical_count': physical_count}
    ]
    
    # 计算资源使用效率
    avg_machines_per_department = round((vm_count + physical_count) / department_count, 2) if department_count > 0 else 0
    avg_machines_per_product = round((vm_count + physical_count) / product_count, 2) if product_count > 0 else 0
    avg_machines_per_dept_product = round((vm_count + physical_count) / (department_count * product_count), 2) if department_count > 0 and product_count > 0 else 0
    
    conn.close()
    
    # 准备统计数据
    stats_data = {
        'vm_count': vm_count,
        'product_count': product_count,
        'physical_count': physical_count,
        'department_count': department_count,
        'group_count': group_count,
        'avg_machines_per_department': avg_machines_per_department,
        'avg_machines_per_product': avg_machines_per_product,
        'avg_machines_per_dept_product': avg_machines_per_dept_product
    }
    
    return render_template('stats.html', active_page='stats', stats=stats_data, product_data=product_data, department_data=department_data, department_product_data=department_product_data, device_type_data=device_type_data, department_product_matrix=department_product_matrix, resource_trend_data=resource_trend_data, resource_quarterly_data=resource_quarterly_data, resource_yearly_data=resource_yearly_data, vm_by_department=vm_by_department, physical_by_department=physical_by_department)

# 检查设备状态API
@app.route('/api/check-device', methods=['POST'])
@login_required
def check_device():
    import subprocess
    import re
    import paramiko
    
    ip = request.form.get('ip')
    username = request.form.get('username')
    password = request.form.get('password')
    
    # 添加日志信息
    print(f"接收到的参数: ip={ip}, username={username}, password={password}")
    
    if not ip:
        return jsonify({'error': '缺少IP地址参数'}), 400
    
    # 初始化返回结果
    result = {
        'ip': ip,
        'status': False,
        'delay': None,
        'cpu_usage': None,
        'memory_usage': None,
        'disk_usage': None,
        'os_type': 'Unknown',
        'ssh_status': 'not_attempted',
        'ssh_error': None
    }
    
    try:
        # 使用ping命令检查设备状态
        import platform
        
        # 根据操作系统选择ping命令参数
        if platform.system() != 'Windows':  # 先判断Linux/macOS
            ping_args = ['ping', '-c', '1', '-W', '1', ip]
        else:  # Windows系统
            ping_args = ['ping', '-n', '1', '-w', '1000', ip]
        
        ping_result = subprocess.run(
            ping_args,
            capture_output=True,
            text=True,
            timeout=5
        )
        
        # 检查ping是否成功
        result['status'] = ping_result.returncode == 0
        
        # 从输出中提取延迟信息
        if result['status']:
            print(f"Ping成功，输出: {ping_result.stdout}")
            # 匹配延迟信息的正则表达式，同时支持Windows和Linux格式
            # Windows格式: 平均 = 4ms 或 时间=4ms
            # Linux格式: time=4ms
            delay_match = re.search(r'(?:time|平均|时间)[=:]\s*(\d+)', ping_result.stdout.lower())
            if delay_match:
                result['delay'] = int(delay_match.group(1))
                print(f"提取到延迟: {result['delay']}ms")
            else:
                print(f"无法从输出中提取延迟信息: {ping_result.stdout}")
        else:
            print(f"Ping失败，返回码: {ping_result.returncode}，输出: {ping_result.stdout}")
        
        # 检查用户名和密码是否为空字符串
        has_valid_credentials = username and password and username.strip() and password.strip()
        
        # 如果ping成功，且提供了有效的用户名和密码，尝试SSH连接获取系统信息
        print(f"准备尝试SSH连接: status={result['status']}, has_valid_credentials={has_valid_credentials}")
        if result['status'] and has_valid_credentials:
            ssh = None
            try:
                print(f"正在建立SSH连接到 {ip}，用户名: {username}")
                # 建立SSH连接
                ssh = paramiko.SSHClient()
                ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                ssh.connect(ip, username=username, password=password, timeout=10)
                print(f"SSH连接成功到 {ip}")
                result['ssh_status'] = 'success'
                
                # 直接尝试执行Linux的SSH指令，不判断设备类型
                # 获取CPU使用率
                print("正在获取CPU使用率")
                try:
                    stdin, stdout, stderr = ssh.exec_command("top -bn1 | head -20 | grep 'Cpu(s)' | awk '{print $2}' | sed 's/%//'")
                    cpu_output = stdout.read().decode().strip()
                    stderr_output = stderr.read().decode().strip()
                    print(f"CPU使用率输出: {cpu_output}")
                    print(f"CPU命令错误输出: {stderr_output}")
                    if cpu_output:
                        result['cpu_usage'] = round(float(cpu_output), 1)
                        print(f"CPU使用率: {result['cpu_usage']}%")
                    else:
                        # 尝试其他命令获取CPU使用率
                        print("尝试使用mpstat命令获取CPU使用率")
                        stdin, stdout, stderr = ssh.exec_command("mpstat 1 1 | tail -1 | awk '{print 100-$12}'")
                        cpu_output = stdout.read().decode().strip()
                        stderr_output = stderr.read().decode().strip()
                        print(f"mpstat输出: {cpu_output}")
                        print(f"mpstat错误输出: {stderr_output}")
                        if cpu_output:
                            result['cpu_usage'] = round(float(cpu_output), 1)
                            print(f"备用CPU使用率: {result['cpu_usage']}%")
                except Exception as e:
                    print(f"获取CPU使用率失败: {str(e)}")
                
                # 获取内存使用率
                print("正在获取内存使用率")
                try:
                    stdin, stdout, stderr = ssh.exec_command("free | grep Mem | awk '{print $3/$2 * 100.0}'")
                    memory_output = stdout.read().decode().strip()
                    stderr_output = stderr.read().decode().strip()
                    print(f"内存使用率输出: {memory_output}")
                    print(f"内存命令错误输出: {stderr_output}")
                    if memory_output:
                        result['memory_usage'] = round(float(memory_output), 1)
                        print(f"内存使用率: {result['memory_usage']}%")
                except Exception as e:
                    print(f"获取内存使用率失败: {str(e)}")
                
                # 获取磁盘使用率
                print("正在获取磁盘使用率")
                try:
                    stdin, stdout, stderr = ssh.exec_command("df -h / | tail -1 | awk '{print $5}' | sed 's/%//'")
                    disk_output = stdout.read().decode().strip()
                    stderr_output = stderr.read().decode().strip()
                    print(f"磁盘使用率输出: {disk_output}")
                    print(f"磁盘命令错误输出: {stderr_output}")
                    if disk_output:
                        result['disk_usage'] = float(disk_output)
                        print(f"磁盘使用率: {result['disk_usage']}%")
                    else:
                        # 尝试其他命令获取磁盘使用率
                        print("尝试使用df -P命令获取磁盘使用率")
                        stdin, stdout, stderr = ssh.exec_command("df -P / | tail -1 | awk '{print $5}' | sed 's/%//'")
                        disk_output = stdout.read().decode().strip()
                        stderr_output = stderr.read().decode().strip()
                        print(f"df -P输出: {disk_output}")
                        print(f"df -P错误输出: {stderr_output}")
                        if disk_output:
                            result['disk_usage'] = float(disk_output)
                            print(f"备用磁盘使用率: {result['disk_usage']}%")
                except Exception as e:
                    print(f"获取磁盘使用率失败: {str(e)}")
            except Exception as ssh_error:
                # SSH连接失败不影响ping结果
                print(f"SSH连接失败: {str(ssh_error)}")
                result['ssh_status'] = 'failed'
                # 优化错误信息显示
                error_message = str(ssh_error)
                if 'Authentication failed' in error_message:
                    result['ssh_error'] = '密码错误'
                elif 'Connection refused' in error_message or 'Unable to connect to port 22' in error_message:
                    result['ssh_error'] = '连接被拒绝（SSH服务未运行或端口未开放）'
                elif 'No route to host' in error_message:
                    result['ssh_error'] = '无路由到主机（网络不可达）'
                elif 'Timeout' in error_message:
                    result['ssh_error'] = '连接超时'
                else:
                    result['ssh_error'] = error_message
            finally:
                # 确保无论是否发生异常，SSH连接都会被关闭
                if ssh is not None:
                    ssh.close()
        else:
            print("跳过SSH连接: 状态为失败或缺少用户名/密码")
            result['ssh_status'] = 'skipped'
            # 根据具体情况返回不同的错误信息
            if not has_valid_credentials:
                result['ssh_error'] = '未获得账号密码'
            else:
                result['ssh_error'] = '网络不可达（设备ping失败）'
        
        return jsonify(result)
        
    except subprocess.TimeoutExpired:
        return jsonify(result)
    except Exception as e:
        return jsonify({
            **result,
            'error': str(e)
        })

# 获取所有设备（用于监控）
@app.route('/api/devices-for-monitor', methods=['GET'])
@login_required
def get_devices_for_monitor():
    conn = get_db_connection()
    devices = conn.execute('SELECT * FROM devices ORDER BY id').fetchall()
    conn.close()
    
    result = []
    for device in devices:
        result.append({
            'id': device['id'],
            'department': device['department'],
            'vm_name': device['vm_name'],
            'ip_address': device['ip_address'],
            'username': device['username'],
            'password': device['password'],
            'remark': device['remark'],
            'created_at': device['created_at'],
            'updated_at': device['updated_at']
        })
    
    return jsonify(result)

# 获取所有物理设备（用于监控）


# 获取所有用户
@app.route('/api/users', methods=['GET'])
@login_required
@admin_required
def get_all_users():
    try:
        conn = get_db_connection()
        users = conn.execute('SELECT users.*, roles.name as role_name FROM users JOIN roles ON users.role_id = roles.id ORDER BY users.id').fetchall()
        conn.close()
        
        result = []
        for user in users:
            result.append({
                'id': user['id'],
                'username': user['username'],
                'role_id': user['role_id'],
                'role_name': user['role_name'],
                'department': user['department'],
                'created_at': user['created_at']
            })
        
        return jsonify(result)
    except Exception as e:
        print(f"获取用户列表失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'获取用户列表失败: {str(e)}'}), 500

# 获取单个用户
@app.route('/api/users/<int:user_id>', methods=['GET'])
@login_required
def get_user(user_id):
    # 获取当前登录用户的ID
    current_user_id = session.get('user_id')
    
    # 检查权限：只有管理员或用户本人才能访问
    user_role = session.get('role')
    if user_role != 'admin' and current_user_id != user_id:
        return jsonify({'error': '权限不足'}), 403
    
    conn = get_db_connection()
    user = conn.execute('SELECT users.*, roles.name as role_name FROM users JOIN roles ON users.role_id = roles.id WHERE users.id = ?', (user_id,)).fetchone()
    conn.close()
    
    if user is None:
        return jsonify({'error': '用户不存在'}), 404
    
    return jsonify({
        'id': user['id'],
        'username': user['username'],
        'role_id': user['role_id'],
        'role_name': user['role_name'],
        'department': user['department'],
        'created_at': user['created_at']
    })

# 获取所有部门
@app.route('/api/departments', methods=['GET'])
@login_required
def get_all_departments():
    try:
        conn = get_db_connection()
        
        # 从devices表获取所有部门
        device_departments = conn.execute('SELECT DISTINCT department FROM devices WHERE department IS NOT NULL AND department != ""').fetchall()
        
        # 从physical_devices表获取所有部门
        physical_departments = conn.execute('SELECT DISTINCT department FROM physical_devices WHERE department IS NOT NULL AND department != ""').fetchall()
        
        conn.close()
        
        # 合并并去重部门列表
        departments = set()
        for dept in device_departments:
            if dept['department']:
                departments.add(dept['department'])
        for dept in physical_departments:
            if dept['department']:
                departments.add(dept['department'])
        
        # 转换为列表并排序
        departments = sorted(list(departments))
        
        return jsonify(departments)
    except Exception as e:
        print(f"获取部门列表失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'获取部门列表失败: {str(e)}'}), 500

# 部门管理API

# 获取所有部门管理记录
@app.route('/api/department-management', methods=['GET'])
@login_required
@admin_required
def get_department_management():
    try:
        print("开始获取部门管理记录...")
        conn = get_db_connection()
        print("数据库连接成功")
        records = conn.execute('SELECT * FROM department_management ORDER BY id').fetchall()
        print(f"查询到 {len(records)} 条记录")
        conn.close()
        result = []
        for record in records:
            result.append({
                'id': record['id'],
                'department': record['department'],
                'device_group': record['device_group'],
                'created_at': record['created_at'],
                'updated_at': record['updated_at']
            })
        print(f"返回 {len(result)} 条记录")
        return jsonify(result)
    except Exception as e:
        import traceback
        print(f"获取部门管理记录失败: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': f'获取记录失败：{str(e)}'}), 500

# 添加部门管理记录
@app.route('/api/department-management', methods=['POST'])
@login_required
@admin_required
def add_department_management():
    try:
        new_record = request.json
        if not new_record or not 'department' in new_record:
            return jsonify({'error': '缺少必要字段'}), 400
        
        department = new_record.get('department', None)
        device_group = new_record.get('device_group', None)
        
        conn = get_db_connection()
        cursor = conn.execute(
            'INSERT INTO department_management (department, device_group) VALUES (?, ?)',
            (department, device_group)
        )
        conn.commit()
        new_id = cursor.lastrowid
        conn.close()
        return jsonify({'message': '记录添加成功', 'id': new_id})
    except Exception as e:
        return jsonify({'error': f'添加失败：{str(e)}'}), 500

# 获取单个部门管理记录
@app.route('/api/department-management/<int:id>', methods=['GET'])
@login_required
@admin_required
def get_department_management_record(id):
    try:
        conn = get_db_connection()
        record = conn.execute('SELECT * FROM department_management WHERE id = ?', (id,)).fetchone()
        conn.close()
        if record is None:
            return jsonify({'error': '记录不存在'}), 404
        return jsonify({
            'id': record['id'],
            'department': record['department'],
            'device_group': record['device_group'],
            'created_at': record['created_at'],
            'updated_at': record['updated_at']
        })
    except Exception as e:
        return jsonify({'error': f'获取记录失败：{str(e)}'}), 500

# 更新部门管理记录
@app.route('/api/department-management/<int:id>', methods=['PUT'])
@login_required
@admin_required
def update_department_management(id):
    try:
        updated_record = request.json
        if not updated_record:
            return jsonify({'error': '缺少更新数据'}), 400
        
        department = updated_record.get('department', None)
        device_group = updated_record.get('device_group', None)
        
        conn = get_db_connection()
        # 检查记录是否存在
        existing_record = conn.execute('SELECT * FROM department_management WHERE id = ?', (id,)).fetchone()
        if existing_record is None:
            conn.close()
            return jsonify({'error': '记录不存在'}), 404
        
        # 获取旧的部门名称和组别名称
        old_department = existing_record['department']
        old_device_group = existing_record['device_group']
        
        # 构建更新语句
        update_fields = []
        update_params = []
        
        if department is not None:
            update_fields.append('department = ?')
            update_params.append(department)
        if device_group is not None:
            update_fields.append('device_group = ?')
            update_params.append(device_group)
        
        if update_fields:
            update_fields.append('updated_at = CURRENT_TIMESTAMP')
            update_sql = 'UPDATE department_management SET ' + ', '.join(update_fields) + ' WHERE id = ?'
            update_params.append(id)
            conn.execute(update_sql, update_params)
            
            # 如果部门名称发生了变化，同时更新所有相关表中的相同部门名称
            if department is not None and department != old_department:
                # 更新部门管理表中所有相同旧部门名称的记录
                conn.execute('UPDATE department_management SET department = ?, updated_at = CURRENT_TIMESTAMP WHERE department = ?', (department, old_department))
                # 更新虚拟机表中的部门名称
                conn.execute('UPDATE devices SET department = ? WHERE department = ?', (department, old_department))
                # 更新物理机表中的部门名称
                conn.execute('UPDATE physical_devices SET department = ? WHERE department = ?', (department, old_department))
            
            # 如果组别名称发生了变化，只更新同一部门下的相同组别名称
            if device_group is not None and device_group != old_device_group:
                # 更新部门管理表中同一部门下的相同旧组别名称的记录
                conn.execute('UPDATE department_management SET device_group = ?, updated_at = CURRENT_TIMESTAMP WHERE device_group = ? AND department = ?', (device_group, old_device_group, old_department))
                # 更新虚拟机表中同一部门下的相同旧组别名称的记录
                conn.execute('UPDATE devices SET device_group = ? WHERE device_group = ? AND department = ?', (device_group, old_device_group, old_department))
            
            conn.commit()
        
        conn.close()
        return jsonify({'message': '记录更新成功'})
    except Exception as e:
        return jsonify({'error': f'更新失败：{str(e)}'}), 500

# 删除部门管理记录
@app.route('/api/department-management/<int:id>', methods=['DELETE'])
@login_required
@admin_required
def delete_department_management(id):
    try:
        conn = get_db_connection()
        # 检查记录是否存在
        existing_record = conn.execute('SELECT * FROM department_management WHERE id = ?', (id,)).fetchone()
        if existing_record is None:
            conn.close()
            return jsonify({'error': '记录不存在'}), 404
        
        conn.execute('DELETE FROM department_management WHERE id = ?', (id,))
        conn.commit()
        conn.close()
        return jsonify({'message': '记录删除成功'})
    except Exception as e:
        return jsonify({'error': f'删除失败：{str(e)}'}), 500

# 导入部门管理记录
@app.route('/api/department-management/import', methods=['POST'])
@login_required
@admin_required
def import_department_management():
    try:
        if 'file' not in request.files:
            return jsonify({'error': '请选择要导入的文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '请选择要导入的文件'}), 400
        
        # 检查文件类型
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': '只支持 .xlsx 和 .xls 格式的Excel文件'}), 400
        
        # 保存临时文件
        temp_file_path = os.path.join(tempfile.gettempdir(), file.filename)
        file.save(temp_file_path)
        
        # 读取Excel文件
        workbook = openpyxl.load_workbook(temp_file_path)
        sheet = workbook.active
        
        # 获取表头
        headers = []
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col).value
            headers.append(header)
        
        # 检查表头
        expected_headers = ['部门', '组别']
        if headers[:2] != expected_headers:
            os.remove(temp_file_path)
            return jsonify({'error': 'Excel文件格式错误，表头应为：部门、组别'}), 400
        
        conn = get_db_connection()
        success_count = 0
        skip_count = 0
        error_count = 0
        errors = []
        
        # 导入数据，将组别字段中的多个团队拆分成单独的记录
        for row in range(2, sheet.max_row + 1):
            department = sheet.cell(row=row, column=1).value
            groups_text = sheet.cell(row=row, column=2).value
            
            # 检查是否是空行
            if not department and not groups_text:
                continue
            
            # 检查必填字段
            if not department:
                error_count += 1
                errors.append(f'第{row}行：部门不能为空')
                continue
            
            # 按空格拆分组别
            if groups_text:
                import re
                groups = re.split(r'\s+', str(groups_text))
                # 过滤空字符串
                groups = [group for group in groups if group.strip()]
                
                # 为每个组别创建一条记录
                for group in groups:
                    try:
                        # 检查是否已存在相同的部门和组别组合
                        existing = conn.execute(
                            'SELECT id FROM department_management WHERE department = ? AND device_group = ?',
                            (department, group)
                        ).fetchone()
                        
                        if existing:
                            # 已存在，跳过
                            skip_count += 1
                        else:
                            # 不存在，插入新记录
                            conn.execute(
                                'INSERT INTO department_management (department, device_group) VALUES (?, ?)',
                                (department, group)
                            )
                            success_count += 1
                    except Exception as e:
                        error_count += 1
                        errors.append(f'第{row}行，组别"{group}"导入失败：{str(e)}')
        
        conn.commit()
        conn.close()
        
        # 删除临时文件
        os.remove(temp_file_path)
        
        return jsonify({
            'message': f'导入完成，成功{success_count}条，跳过{skip_count}条，失败{error_count}条',
            'success_count': success_count,
            'skip_count': skip_count,
            'error_count': error_count,
            'errors': errors
        })
    except Exception as e:
        return jsonify({'error': f'导入失败：{str(e)}'}), 500

# 重置部门管理数据
@app.route('/api/department-management/reset', methods=['POST'])
@login_required
@admin_required
def reset_department_management():
    try:
        conn = get_db_connection()
        
        # 清空所有部门管理数据并重置自增ID
        cursor = conn.execute('DELETE FROM department_management')
        deleted_count = cursor.rowcount
        
        # 重置自增ID计数器
        conn.execute('DELETE FROM sqlite_sequence WHERE name="department_management"')
          
        conn.commit()
        conn.close()
        
        return jsonify({'message': f'数据清空成功，共删除{deleted_count}条记录，ID已重置'})
    except Exception as e:
        return jsonify({'error': f'清空失败：{str(e)}'}), 500

# 添加用户
@app.route('/api/users', methods=['POST'])
@login_required
@admin_required
def add_user():
    try:
        new_user = request.get_json()
        
        if not all(key in new_user for key in ['username', 'password', 'role_id']):
            return jsonify({'error': '缺少必要字段'}), 400

        # 获取部门字段，如果不存在则设为None
        department = new_user.get('department', None)

        conn = get_db_connection()
        cursor = conn.execute(
            'INSERT INTO users (username, password, role_id, department) VALUES (?, ?, ?, ?)',
            (new_user['username'], new_user['password'], new_user['role_id'], department)
        )
        conn.commit()
        new_id = cursor.lastrowid
        return jsonify({'message': '用户添加成功', 'id': new_id})  
    except sqlite3.IntegrityError:
        return jsonify({'error': '用户名已存在'}), 409
    except Exception as e:
        print(f"添加用户失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'添加用户失败: {str(e)}'}), 500
    finally:
        conn.close()

# 更新用户
@app.route('/api/users/<int:user_id>', methods=['PUT'])
@login_required
@admin_required
def update_user(user_id):
    try:
        updated_data = request.get_json()
        conn = get_db_connection()
        
        # 检查用户是否存在
        user = conn.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
        if user is None:
            conn.close()
            return jsonify({'error': '用户不存在'}), 404
        
        # 更新用户信息
        update_fields = []
        update_values = []
        
        if 'username' in updated_data:
            update_fields.append("username = ?")
            update_values.append(updated_data['username'])
        
        if 'password' in updated_data and updated_data['password']:
            update_fields.append("password = ?")
            update_values.append(updated_data['password'])
        
        if 'role_id' in updated_data:
            update_fields.append("role_id = ?")
            update_values.append(updated_data['role_id'])
        
        if 'department' in updated_data:
            update_fields.append("department = ?")
            update_values.append(updated_data['department'])
        
        if not update_fields:
            conn.close()
            return jsonify({'error': '没有需要更新的字段'}), 400
        
        update_values.append(user_id)
        
        conn.execute(
            f'UPDATE users SET {" , ".join(update_fields)} WHERE id = ?',
            update_values
        )
        conn.commit()
        return jsonify({'message': '用户更新成功'})
    except sqlite3.IntegrityError:
        return jsonify({'error': '用户名已存在'}), 409
    except Exception as e:
        print(f"更新用户失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'更新用户失败: {str(e)}'}), 500
    finally:
        conn.close()

# 删除用户
@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_user(user_id):
    try:
        conn = get_db_connection()
        
        # 检查用户是否存在
        user = conn.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
        if user is None:
            conn.close()
            return jsonify({'error': '用户不存在'}), 404
        
        # 检查是否是admin用户
        if user['username'] == 'admin':
            conn.close()
            return jsonify({'error': '不能删除admin用户'}), 403
        
        conn.execute('DELETE FROM users WHERE id = ?', (user_id,))
        conn.commit()
        conn.close()
        
        return jsonify({'message': '用户删除成功', 'id': user_id})
    except Exception as e:
        print(f"删除用户失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'删除用户失败: {str(e)}'}), 500

# Dash应用布局
dash_app.layout = html.Div([
    html.H1('设备数据可视化', style={'text-align': 'center', 'margin-bottom': '30px'}),
    
    # 部门选择器
    html.Div([
        html.Label('选择部门:', style={'margin-right': '10px'}),
        dcc.Dropdown(
            id='department-dropdown',
            options=[],  # 动态加载部门选项
            value='all',
            placeholder='选择部门...',
            multi=True
        )
    ], style={'margin-bottom': '20px'}),
    
    # 图表区域
    html.Div([
        # 第一个图表：部门设备分布（饼图）
        html.Div([
            dcc.Graph(id='department-chart')
        ], style={'width': '48%', 'display': 'inline-block', 'margin-right': '4%'}),
        
        # 第二个图表：产品设备分布（柱状图）
        html.Div([
            dcc.Graph(id='product-chart')
        ], style={'width': '48%', 'display': 'inline-block'})
    ], style={'margin-bottom': '30px'}),
    
    # 第三个图表：设备类型分布（折线图）
    html.Div([
        dcc.Graph(id='device-type-chart')
    ])
])

# 加载部门选项的回调
@dash_app.callback(
    Output('department-dropdown', 'options'),
    Input('department-dropdown', 'value')
)
def load_departments(value):
    conn = get_db_connection()
    # 从设备表获取所有部门
    departments = conn.execute('SELECT DISTINCT department FROM devices WHERE department IS NOT NULL AND department != ""').fetchall()
    # 从物理设备表获取所有部门
    physical_departments = conn.execute('SELECT DISTINCT department FROM physical_devices WHERE department IS NOT NULL AND department != ""').fetchall()
    conn.close()
    
    # 合并并去重部门列表
    dept_set = set()
    for dept in departments:
        if dept['department']:
            dept_set.add(dept['department'])
    for dept in physical_departments:
        if dept['department']:
            dept_set.add(dept['department'])
    
    # 转换为选项格式
    options = [{'label': dept, 'value': dept} for dept in sorted(list(dept_set))]
    options.insert(0, {'label': '全部部门', 'value': 'all'})
    
    return options

# 更新部门分布图表的回调
@dash_app.callback(
    Output('department-chart', 'figure'),
    Input('department-dropdown', 'value')
)
def update_department_chart(selected_depts):
    conn = get_db_connection()
    
    # 构建查询条件
    if selected_depts and 'all' not in selected_depts:
        placeholders = ','.join(['?'] * len(selected_depts))
        query = f'SELECT department, COUNT(*) as count FROM devices WHERE department IN ({placeholders}) GROUP BY department'
        devices = conn.execute(query, selected_depts).fetchall()
    else:
        query = 'SELECT department, COUNT(*) as count FROM devices GROUP BY department'
        devices = conn.execute(query).fetchall()
    
    conn.close()
    
    # 转换为DataFrame
    df = pd.DataFrame(devices, columns=['department', 'count'])
    
    # 创建饼图
    fig = px.pie(df, values='count', names='department', title='部门设备分布')
    fig.update_layout(transition_duration=500)
    
    return fig

# 更新产品分布图表的回调
@dash_app.callback(
    Output('product-chart', 'figure'),
    Input('department-dropdown', 'value')
)
def update_product_chart(selected_depts):
    conn = get_db_connection()
    
    # 构建查询条件
    if selected_depts and 'all' not in selected_depts:
        placeholders = ','.join(['?'] * len(selected_depts))
        query = f'SELECT product, COUNT(*) as count FROM devices WHERE department IN ({placeholders}) GROUP BY product'
        devices = conn.execute(query, selected_depts).fetchall()
    else:
        query = 'SELECT product, COUNT(*) as count FROM devices GROUP BY product'
        devices = conn.execute(query).fetchall()
    
    conn.close()
    
    # 转换为DataFrame
    df = pd.DataFrame(devices, columns=['product', 'count'])
    
    # 创建柱状图
    fig = px.bar(df, x='product', y='count', title='产品设备分布')
    fig.update_layout(transition_duration=500)
    
    return fig

# 更新设备类型分布图表的回调
@dash_app.callback(
    Output('device-type-chart', 'figure'),
    Input('department-dropdown', 'value')
)
def update_device_type_chart(selected_depts):
    conn = get_db_connection()
    
    # 获取虚拟机设备数量
    if selected_depts and 'all' not in selected_depts:
        placeholders = ','.join(['?'] * len(selected_depts))
        vm_query = f'SELECT COUNT(*) as count FROM devices WHERE department IN ({placeholders})'
        vm_count = conn.execute(vm_query, selected_depts).fetchone()[0]
        
        # 获取物理机设备数量
        physical_query = f'SELECT COUNT(*) as count FROM physical_devices WHERE department IN ({placeholders})'
        physical_count = conn.execute(physical_query, selected_depts).fetchone()[0]
    else:
        vm_count = conn.execute('SELECT COUNT(*) as count FROM devices').fetchone()[0]
        physical_count = conn.execute('SELECT COUNT(*) as count FROM physical_devices').fetchone()[0]
    
    conn.close()
    
    # 转换为DataFrame
    df = pd.DataFrame([
        {'type': '虚拟机', 'count': vm_count},
        {'type': '物理机', 'count': physical_count}
    ])
    
    # 创建折线图
    fig = px.line(df, x='type', y='count', title='设备类型分布')
    fig.update_layout(transition_duration=500)
    
    return fig

if __name__ == '__main__':
    # 获取应用的绝对路径
    import os
    app_root = os.path.dirname(os.path.abspath(__file__))
    
    # 切换到应用根目录
    os.chdir(app_root)
    print(f"切换工作目录到: {app_root}")
    
    # 启动应用
    app.run(debug=True, host='0.0.0.0', port=5000)