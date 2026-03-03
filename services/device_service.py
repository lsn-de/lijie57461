from utils.db import DatabaseManager
import openpyxl


class DeviceService:
    """设备服务"""
    
    @staticmethod
    def get_devices(product=None, department=None, group=None):
        """获取设备列表"""
        # 构建查询条件
        conditions = []
        params = []
        
        if product:
            conditions.append("product LIKE ?")
            params.append(f"%{product}%")
        if department:
            conditions.append("department LIKE ?")
            params.append(f"%{department}%")
        if group:
            conditions.append("device_group LIKE ?")
            params.append(f"%{group}%")
        
        # 构建SQL查询
        if conditions:
            # 有筛选参数，使用模糊匹配
            query = 'SELECT * FROM devices WHERE ' + ' AND '.join(conditions) + ' ORDER BY id'
            devices = DatabaseManager.execute_query(query, params, fetchall=True)
        else:
            # 无筛选参数，获取所有设备
            devices = DatabaseManager.execute_query('SELECT * FROM devices ORDER BY id', fetchall=True)
        
        return [dict(device) for device in devices]
    
    @staticmethod
    def get_device_by_id(device_id):
        """根据ID获取设备"""
        device = DatabaseManager.execute_query('SELECT * FROM devices WHERE id = ?', (device_id,), fetchone=True)
        return dict(device) if device else None
    
    @staticmethod
    def add_device(product, department, group, vm_name, ip_address, username, password, remark=None):
        """添加设备"""
        try:
            # 插入数据
            DatabaseManager.execute_query(
                'INSERT INTO devices (product, department, device_group, vm_name, ip_address, username, password, remark) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
                (product, department, group, vm_name, ip_address, username, password, remark)
            )
            # 获取新插入记录的ID
            new_id = DatabaseManager.execute_query('SELECT last_insert_rowid()', fetchone=True)[0]
            return new_id
        except Exception as e:
            raise e
    
    @staticmethod
    def update_device(device_id, product, department, group, vm_name, ip_address, username, password, remark=None):
        """更新设备"""
        try:
            DatabaseManager.execute_query('''
                UPDATE devices
                SET product = ?, department = ?, device_group = ?, vm_name = ?, ip_address = ?, username = ?, password = ?, remark = ?
                WHERE id = ?
            ''', (
                product, department, group, vm_name, ip_address, username, password, remark, device_id
            ))
            return True
        except Exception as e:
            raise e
    
    @staticmethod
    def delete_device(device_id):
        """删除设备"""
        try:
            DatabaseManager.execute_query('DELETE FROM devices WHERE id = ?', (device_id,))
            return True
        except Exception as e:
            raise e
    
    @staticmethod
    def import_devices(file):
        """从Excel导入设备数据"""
        try:
            # 读取Excel文件
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
            
            # 检查表头
            headers = []
            for col in range(1, sheet.max_column + 1):
                header = sheet.cell(row=1, column=col).value
                headers.append(header)
            
            expected_headers = ['产品', '部门', '组别', '虚拟机名称', 'IP地址', '账号', '密码']
            # 允许第八列为备注字段
            if headers[:7] != expected_headers:
                raise ValueError('Excel文件格式错误，表头应为：产品、部门、组别、虚拟机名称、IP地址、账号、密码')
            
            success_count = 0
            error_count = 0
            errors = []
            
            # 导入数据
            for row in range(2, sheet.max_row + 1):
                product = sheet.cell(row=row, column=1).value
                department = sheet.cell(row=row, column=2).value
                group = sheet.cell(row=row, column=3).value
                vm_name = sheet.cell(row=row, column=4).value
                ip_address = sheet.cell(row=row, column=5).value
                username = sheet.cell(row=row, column=6).value
                password = sheet.cell(row=row, column=7).value
                
                # 读取备注字段（如果存在）
                remark = sheet.cell(row=row, column=8).value if sheet.max_column >= 8 else None
                
                # 检查是否是空行（所有必填字段都为空）
                if not vm_name and not ip_address and not username and not password:
                    # 空行，跳过处理
                    continue
                
                # 详细检查每个必填字段
                missing_fields = []
                if not vm_name:
                    missing_fields.append('虚拟机名称')
                if not ip_address:
                    missing_fields.append('IP地址')
                if not username:
                    missing_fields.append('账号')
                if not password:
                    missing_fields.append('密码')
                
                if missing_fields:
                    error_count += 1
                    errors.append(f'第{row}行数据不完整：缺少{"、".join(missing_fields)}')
                    continue
                
                try:
                    DeviceService.add_device(product, department, group, vm_name, ip_address, username, password, remark)
                    success_count += 1
                except Exception as e:
                    error_count += 1
                    if 'ip_address' in str(e):
                        errors.append(f'第{row}行IP地址({ip_address})已存在')
                    else:
                        errors.append(f'第{row}行数据处理错误：{str(e)}')
            
            return {
                'success_count': success_count,
                'error_count': error_count,
                'errors': errors
            }
        except Exception as e:
            raise e
    
    @staticmethod
    def reset_database():
        """重置数据库"""
        try:
            # 清空devices表
            DatabaseManager.execute_query('DELETE FROM devices')
            
            # 重置自增ID
            DatabaseManager.execute_query('DELETE FROM sqlite_sequence WHERE name = "devices"')
            
            return True
        except Exception as e:
            raise e
    
    @staticmethod
    def get_all_vms():
        """获取所有虚拟机列表"""
        vms = DatabaseManager.execute_query('SELECT id, ip_address, username, password FROM devices', fetchall=True)
        return [dict(vm) for vm in vms]
