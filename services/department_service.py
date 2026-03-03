from utils.db import DatabaseManager
import openpyxl


class DepartmentService:
    """部门服务"""
    
    @staticmethod
    def get_department_records(page=1, page_size=10, department=None, device_group=None):
        """获取部门管理记录"""
        # 构建查询条件
        conditions = []
        params = []
        
        if department:
            conditions.append("department LIKE ?")
            params.append(f"%{department}%")
        if device_group:
            conditions.append("device_group LIKE ?")
            params.append(f"%{device_group}%")
        
        # 构建SQL查询
        sql = "SELECT * FROM department_management"
        if conditions:
            sql += " WHERE " + " AND ".join(conditions)
        
        # 获取总记录数
        count_sql = "SELECT COUNT(*) FROM department_management"
        if conditions:
            count_sql += " WHERE " + " AND ".join(conditions)
        total_records = DatabaseManager.execute_query(count_sql, params, fetchone=True)[0]
        
        # 分页查询
        offset = (page - 1) * page_size
        sql += " ORDER BY id DESC LIMIT ? OFFSET ?"
        params.extend([page_size, offset])
        
        records = DatabaseManager.execute_query(sql, params, fetchall=True)
        records_list = [dict(record) for record in records]
        
        return {
            'records': records_list,
            'total': total_records,
            'page': page,
            'page_size': page_size
        }
    
    @staticmethod
    def add_department_record(department, device_group):
        """添加部门管理记录"""
        if not department or not device_group:
            raise ValueError('部门和组别不能为空')
        
        try:
            DatabaseManager.execute_query(
                'INSERT INTO department_management (department, device_group) VALUES (?, ?)',
                (department, device_group)
            )
            return True
        except Exception as e:
            raise e
    
    @staticmethod
    def update_department_record(record_id, department, device_group):
        """更新部门管理记录"""
        if not department or not device_group:
            raise ValueError('部门和组别不能为空')
        
        try:
            DatabaseManager.execute_query(
                'UPDATE department_management SET department = ?, device_group = ? WHERE id = ?',
                (department, device_group, record_id)
            )
            return True
        except Exception as e:
            raise e
    
    @staticmethod
    def delete_department_record(record_id):
        """删除部门管理记录"""
        try:
            DatabaseManager.execute_query('DELETE FROM department_management WHERE id = ?', (record_id,))
            return True
        except Exception as e:
            raise e
    
    @staticmethod
    def import_department_records(file):
        """导入部门管理记录"""
        if not file.filename.endswith('.xlsx'):
            raise ValueError('只支持xlsx格式的文件')
        
        try:
            # 读取Excel文件
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
            
            # 检查列数
            if sheet.max_column < 2:
                raise ValueError('Excel文件格式错误：至少需要2列数据（部门和组别）')
            
            success_count = 0
            skipped_count = 0
            
            # 跳过表头行（第一行）
            for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not row[0] or not row[1]:  # 如果部门或组别为空，则跳过该行
                        skipped_count += 1
                        continue
                    
                    # 插入数据到数据库
                    DepartmentService.add_department_record(row[0], row[1])
                    
                    success_count += 1
                except Exception as e:
                    skipped_count += 1
            
            return {
                'success_count': success_count,
                'skipped_count': skipped_count
            }
        except Exception as e:
            raise e
