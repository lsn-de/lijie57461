from utils.db import DatabaseManager
import openpyxl
import logging


class PhysicalService:
    """物理机服务"""
    
    @staticmethod
    def get_physical_devices(product=None, department=None, device_code=None, ip_address=None, oob_ip=None, account=None, password=None, remark=None):
        """获取物理机列表"""
        try:
            logging.debug('开始处理获取物理机请求')
            
            # 构建查询条件
            conditions = []
            params = []
            
            if product:
                conditions.append("product LIKE ?")
                params.append(f"%{product}%")
            if department:
                conditions.append("department LIKE ?")
                params.append(f"%{department}%")
            if device_code:
                conditions.append("device_code LIKE ?")
                params.append(f"%{device_code}%")
            if ip_address:
                conditions.append("ip_address LIKE ?")
                params.append(f"%{ip_address}%")
            if oob_ip:
                conditions.append("oob_ip LIKE ?")
                params.append(f"%{oob_ip}%")
            if account:
                conditions.append("account LIKE ?")
                params.append(f"%{account}%")
            if password:
                conditions.append("password LIKE ?")
                params.append(f"%{password}%")
            if remark:
                conditions.append("remark LIKE ?")
                params.append(f"%{remark}%")
            
            # 构建SQL查询
            sql = "SELECT * FROM physical_devices"
            if conditions:
                sql += " WHERE " + " AND ".join(conditions)
            sql += " ORDER BY id"
            
            logging.debug(f'执行SQL查询: {sql}, 参数: {params}')
            
            # 执行查询
            devices = DatabaseManager.execute_query(sql, params, fetchall=True)
            logging.debug(f'查询结果数量: {len(devices)}')
            
            return [dict(device) for device in devices]
        except Exception as e:
            logging.error(f'获取物理机数据失败: {str(e)}')
            import traceback
            logging.error(traceback.format_exc())
            raise e
    
    @staticmethod
    def get_physical_device_by_id(device_id):
        """根据ID获取物理机"""
        device = DatabaseManager.execute_query('SELECT * FROM physical_devices WHERE id = ?', (device_id,), fetchone=True)
        return dict(device) if device else None
    
    @staticmethod
    def add_physical_device(product, department, device_code, ip_address, oob_ip, account, password, remark=None):
        """添加物理机"""
        try:
            DatabaseManager.execute_query('''
                INSERT INTO physical_devices (product, department, device_code, ip_address, oob_ip, account, password, remark)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                product, department, device_code, ip_address, oob_ip, account, password, remark
            ))
            return True
        except Exception as e:
            raise e
    
    @staticmethod
    def update_physical_device(device_id, product, department, device_code, ip_address, oob_ip, account, password, remark=None):
        """更新物理机"""
        try:
            DatabaseManager.execute_query('''
                UPDATE physical_devices
                SET product = ?, department = ?, device_code = ?, ip_address = ?, oob_ip = ?, account = ?, password = ?, remark = ?
                WHERE id = ?
            ''', (
                product, department, device_code, ip_address, oob_ip, account, password, remark, device_id
            ))
            return True
        except Exception as e:
            raise e
    
    @staticmethod
    def delete_physical_device(device_id):
        """删除物理机"""
        try:
            DatabaseManager.execute_query('DELETE FROM physical_devices WHERE id = ?', (device_id,))
            return True
        except Exception as e:
            raise e
    
    @staticmethod
    def import_physical_devices(file):
        """导入物理机"""
        logging.debug('开始处理物理机导入请求')
        try:
            # 检查文件类型
            if not file.filename.endswith('.xlsx'):
                raise ValueError('只支持xlsx格式的文件')
            
            # 读取Excel文件
            logging.debug('开始读取Excel文件')
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
            logging.debug(f'成功加载Excel文件，工作表：{sheet.title}')
            
            # 计算最大列数，确保行的完整性
            max_columns = sheet.max_column
            logging.debug(f'Excel文件最大列数：{max_columns}')
            
            if max_columns < 8:
                raise ValueError(f'Excel文件格式错误：至少需要8列数据，但只找到{max_columns}列')
            
            # 跳过表头行（第一行）
            success_count = 0
            skipped_count = 0
            
            logging.debug('开始处理Excel数据行')
            for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                logging.debug(f'处理第{row_index}行数据')
                try:
                    # 详细检查每个必填字段
                    product = row[0]
                    department = row[1]
                    device_code = row[2]
                    ip_address = row[3]
                    oob_ip = row[4]
                    account = row[5]
                    password = row[6]
                    remark = row[7] if len(row) > 7 else None
                    
                    missing_fields = []
                    if not department:
                        missing_fields.append('部门')
                    if not ip_address:
                        missing_fields.append('IP地址')
                    if not oob_ip:
                        missing_fields.append('带外IP')
                    if not account:
                        missing_fields.append('账号')
                    if not password:
                        missing_fields.append('密码')
                    
                    if missing_fields:
                        logging.debug(f'第{row_index}行：必填字段不完整，跳过')
                        skipped_count += 1
                        continue
                    
                    # 插入数据到数据库
                    logging.debug(f'第{row_index}行：准备插入数据库')
                    PhysicalService.add_physical_device(
                        product, department, device_code, ip_address, oob_ip, account, password, remark
                    )
                    
                    success_count += 1
                    logging.debug(f'第{row_index}行：数据插入成功')
                    
                except Exception as e:
                    logging.error(f'第{row_index}行：数据处理错误 - {str(e)}')
                    raise e
            
            logging.debug(f'数据行处理完成，成功：{success_count}，跳过：{skipped_count}')
            return {
                'success_count': success_count,
                'skipped_count': skipped_count
            }
        except Exception as e:
            logging.error(f'导入失败：{str(e)}')
            import traceback
            logging.error(traceback.format_exc())
            raise e
    
    @staticmethod
    def reset_physical_database():
        """重置物理机数据库"""
        try:
            # 清空physical_devices表
            DatabaseManager.execute_query('DELETE FROM physical_devices')
            
            # 重置自增ID
            DatabaseManager.execute_query('DELETE FROM sqlite_sequence WHERE name = "physical_devices"')
            
            return True
        except Exception as e:
            raise e
    
    @staticmethod
    def get_all_physical_devices():
        """获取所有物理机列表"""
        devices = DatabaseManager.execute_query('SELECT id, ip_address, account, password FROM physical_devices', fetchall=True)
        return [dict(device) for device in devices]
