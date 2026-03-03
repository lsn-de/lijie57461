from flask import Blueprint, request, jsonify, render_template, flash, redirect, url_for, session
from utils.db import DatabaseManager
from utils.decorators import login_required, admin_required
import openpyxl

# 创建部门管理蓝图
department_bp = Blueprint('department', __name__)


@department_bp.route('/department')
@login_required
def department():
    """部门管理页面"""
    # 仅管理员可见
    if session.get('role') != 'admin':
        flash('无权限访问此页面', 'danger')
        return redirect(url_for('index'))
    return render_template('department.html')


@department_bp.route('/api/department-records', methods=['GET'])
@login_required
def get_department_records():
    """获取部门管理记录"""
    # 仅管理员可见
    if session.get('role') != 'admin':
        return jsonify({'error': '权限不足，需要管理员权限'}), 403
    
    # 获取查询参数
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 10, type=int)
    department = request.args.get('department')
    device_group = request.args.get('device_group')
    
    # 构建查询条件
    conditions = []
    params = []
    
    if department:
        conditions.append("department LIKE ?")
        params.append(f"%{department}%")
    if device_group:
        conditions.append("device_group LIKE ?")
        params.append(f"%{device_group}%")
    
    # 构建SQL查询
    sql = "SELECT * FROM department_management"
    if conditions:
        sql += " WHERE " + " AND ".join(conditions)
    
    # 获取总记录数
    count_sql = "SELECT COUNT(*) FROM department_management"
    if conditions:
        count_sql += " WHERE " + " AND ".join(conditions)
    total_records = DatabaseManager.execute_query(count_sql, params, fetchone=True)[0]
    
    # 分页查询
    offset = (page - 1) * page_size
    sql += " ORDER BY id DESC LIMIT ? OFFSET ?"
    params.extend([page_size, offset])
    
    records = DatabaseManager.execute_query(sql, params, fetchall=True)
    records_list = [dict(record) for record in records]
    
    return jsonify({
        'records': records_list,
        'total': total_records,
        'page': page,
        'page_size': page_size
    })


@department_bp.route('/api/department-records', methods=['POST'])
@login_required
def add_department_record():
    """添加部门管理记录"""
    # 仅管理员可见
    if session.get('role') != 'admin':
        return jsonify({'error': '权限不足，需要管理员权限'}), 403
    
    data = request.form
    department = data.get('department')
    device_group = data.get('device_group')
    
    if not department or not device_group:
        return jsonify({'error': '部门和组别不能为空'}), 400
    
    try:
        DatabaseManager.execute_query(
            'INSERT INTO department_management (department, device_group) VALUES (?, ?)',
            (department, device_group)
        )
        return jsonify({'message': '部门管理记录添加成功'})
    except Exception as e:
        return jsonify({'error': f'添加失败：{str(e)}'}), 500


@department_bp.route('/api/department-records/<int:record_id>', methods=['PUT'])
@login_required
def update_department_record(record_id):
    """更新部门管理记录"""
    # 仅管理员可见
    if session.get('role') != 'admin':
        return jsonify({'error': '权限不足，需要管理员权限'}), 403
    
    data = request.form
    department = data.get('department')
    device_group = data.get('device_group')
    
    if not department or not device_group:
        return jsonify({'error': '部门和组别不能为空'}), 400
    
    try:
        DatabaseManager.execute_query(
            'UPDATE department_management SET department = ?, device_group = ? WHERE id = ?',
            (department, device_group, record_id)
        )
        return jsonify({'message': '部门管理记录更新成功'})
    except Exception as e:
        return jsonify({'error': f'更新失败：{str(e)}'}), 500


@department_bp.route('/api/department-records/<int:record_id>', methods=['DELETE'])
@login_required
def delete_department_record(record_id):
    """删除部门管理记录"""
    # 仅管理员可见
    if session.get('role') != 'admin':
        return jsonify({'error': '权限不足，需要管理员权限'}), 403
    
    try:
        DatabaseManager.execute_query('DELETE FROM department_management WHERE id = ?', (record_id,))
        return jsonify({'message': '部门管理记录删除成功'})
    except Exception as e:
        return jsonify({'error': f'删除失败：{str(e)}'}), 500


@department_bp.route('/api/department-records/import', methods=['POST'])
@login_required
def import_department_records():
    """导入部门管理记录"""
    # 仅管理员可见
    if session.get('role') != 'admin':
        return jsonify({'error': '权限不足，需要管理员权限'}), 403
    
    if 'file' not in request.files:
        return jsonify({'error': '没有文件上传'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '没有选择文件'}), 400
    
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': '只支持xlsx格式的文件'}), 400
    
    try:
        # 读取Excel文件
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        
        # 检查列数
        if sheet.max_column < 2:
            return jsonify({'error': 'Excel文件格式错误：至少需要2列数据（部门和组别）'}), 400
        
        success_count = 0
        skipped_count = 0
        
        # 跳过表头行（第一行）
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row[0] or not row[1]:  # 如果部门或组别为空，则跳过该行
                    skipped_count += 1
                    continue
                
                # 插入数据到数据库
                DatabaseManager.execute_query(
                    'INSERT INTO department_management (department, device_group) VALUES (?, ?)',
                    (row[0], row[1])
                )
                
                success_count += 1
            except Exception as e:
                skipped_count += 1
        
        return jsonify({
            'message': '部门管理记录导入成功',
            'success_count': success_count,
            'skipped_count': skipped_count
        })
    except Exception as e:
        return jsonify({'error': f'导入失败：{str(e)}'}), 500
