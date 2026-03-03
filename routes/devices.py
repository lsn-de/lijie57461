from flask import Blueprint, request, jsonify, session
from utils.decorators import login_required, admin_required
import openpyxl
import logging

# 创建设备管理蓝图
devices_bp = Blueprint('devices', __name__)


@devices_bp.route('/api/devices', methods=['GET'])
@login_required
def get_devices():
    """获取所有设备"""
    # 导入数据库连接函数
    from app import get_db_connection
    conn = get_db_connection()
    try:
        # 获取查询参数
        product = request.args.get('product')
        department = request.args.get('department')
        group = request.args.get('group')
        
        # 构建查询条件
        conditions = []
        params = []
        
        if product:
            conditions.append("product LIKE ?")
            params.append(f"%{product}%")
        if department:
            conditions.append("department LIKE ?")
            params.append(f"%{department}%")
        if group:
            conditions.append("device_group LIKE ?")
            params.append(f"%{group}%")
        
        # 构建SQL查询
        if conditions:
            # 有筛选参数，使用模糊匹配
            query = 'SELECT * FROM devices WHERE ' + ' AND '.join(conditions) + ' ORDER BY id'
            devices = conn.execute(query, params).fetchall()
        else:
            # 无筛选参数，获取所有设备
            devices = conn.execute('SELECT * FROM devices ORDER BY id').fetchall()
        
        # 获取当前登录用户的信息
        user_id = session.get('user_id')
        user = conn.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
        user_role = session.get('role')
        user_department = user['department'] if user else None
        
        result = []
        for device in devices:
            device_dict = {
                'id': device['id'],
                'product': device['product'],
                'department': device['department'],
                'group': device['device_group'],
                'vm_name': device['vm_name'],
                'ip_address': device['ip_address'],
                'username': device['username'],
                'remark': device['remark'],
                'created_at': device['created_at'],
                'updated_at': device['updated_at']
            }
            
            # 只有管理员或同部门用户才能看到密码
            if user_role == 'admin' or (user_department and device['department'] == user_department):
                device_dict['password'] = device['password']
            
            result.append(device_dict)
        
        return jsonify(result)
    finally:
        conn.close()


@devices_bp.route('/api/devices/<int:device_id>', methods=['GET'])
@login_required
def get_device(device_id):
    """获取单个设备"""
    # 导入数据库连接函数
    from app import get_db_connection
    conn = get_db_connection()
    try:
        device = conn.execute('SELECT * FROM devices WHERE id = ?', (device_id,)).fetchone()
        
        # 获取当前登录用户的信息
        user_id = session.get('user_id')
        user = conn.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
        user_role = session.get('role')
        user_department = user['department'] if user else None
        
        if device is None:
            return jsonify({'error': '设备不存在'}), 404
        
        device_dict = {
            'id': device['id'],
            'product': device['product'],
            'department': device['department'],
            'group': device['device_group'],
            'vm_name': device['vm_name'],
            'ip_address': device['ip_address'],
            'username': device['username'],
            'remark': device['remark'],
            'created_at': device['created_at'],
            'updated_at': device['updated_at']
        }
        
        # 只有管理员或同部门用户才能看到密码
        if user_role == 'admin' or (user_department and device['department'] == user_department):
            device_dict['password'] = device['password']
        
        return jsonify(device_dict)
    finally:
        conn.close()


@devices_bp.route('/api/devices', methods=['POST'])
@login_required
def add_device():
    """添加设备"""
    # 支持 JSON 和表单数据
    if request.is_json:
        data = request.get_json()
    else:
        data = request.form
    
    if not all(key in data for key in ['vm_name', 'ip_address', 'username', 'password']):
        return jsonify({'error': '缺少必要字段'}), 400

    # 获取产品、部门和备注字段，如果不存在则设为None
    product = data.get('product') or None
    department = data.get('department') or None
    group = data.get('group') or None
    remark = data.get('remark') or None

    # 导入数据库连接函数
    from app import get_db_connection
    conn = get_db_connection()
    try:
        # 执行插入操作
        cursor = conn.execute('''
            INSERT INTO devices (product, department, device_group, vm_name, ip_address, username, password, remark)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            product, department, group, data['vm_name'], data['ip_address'], data['username'], data['password'], remark
        ))
        
        # 获取新插入记录的ID
        new_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        conn.commit()
        return jsonify({'message': '设备添加成功', 'id': new_id})
    except Exception as e:
        return jsonify({'error': f'添加失败：{str(e)}'}), 500
    finally:
        conn.close()


@devices_bp.route('/api/devices/<int:device_id>', methods=['PUT'])
@login_required
def update_device(device_id):
    """更新设备"""
    # 支持 JSON 和表单数据
    if request.is_json:
        data = request.get_json()
    else:
        data = request.form
    
    # 导入数据库连接函数
    from app import get_db_connection
    conn = get_db_connection()
    try:
        # 检查设备是否存在
        device = conn.execute('SELECT * FROM devices WHERE id = ?', (device_id,)).fetchone()
        if device is None:
            return jsonify({'error': '设备不存在'}), 404
        
        # 构建更新语句，只更新提供的字段
        update_fields = []
        update_params = []
        
        if 'product' in data:
            update_fields.append('product = ?')
            update_params.append(data['product'])
        if 'department' in data:
            update_fields.append('department = ?')
            update_params.append(data['department'])
        if 'group' in data:
            update_fields.append('device_group = ?')
            update_params.append(data['group'])
        if 'vm_name' in data:
            update_fields.append('vm_name = ?')
            update_params.append(data['vm_name'])
        if 'ip_address' in data:
            update_fields.append('ip_address = ?')
            update_params.append(data['ip_address'])
        if 'username' in data:
            update_fields.append('username = ?')
            update_params.append(data['username'])
        if 'password' in data:
            update_fields.append('password = ?')
            update_params.append(data['password'])
        if 'remark' in data:
            update_fields.append('remark = ?')
            update_params.append(data['remark'])
        
        # 如果没有字段需要更新，直接返回成功
        if not update_fields:
            return jsonify({'message': '设备更新成功'})
        
        # 构建 SQL 语句
        sql = f"UPDATE devices SET {', '.join(update_fields)} WHERE id = ?"
        update_params.append(device_id)
        
        conn.execute(sql, update_params)
        conn.commit()
        return jsonify({'message': '设备更新成功'})
    except Exception as e:
        return jsonify({'error': f'更新失败：{str(e)}'}), 500
    finally:
        conn.close()


@devices_bp.route('/api/devices/<int:device_id>', methods=['DELETE'])
@admin_required
def delete_device(device_id):
    """删除设备"""
    # 导入数据库连接函数
    from app import get_db_connection
    conn = get_db_connection()
    try:
        # 检查设备是否存在
        device = conn.execute('SELECT * FROM devices WHERE id = ?', (device_id,)).fetchone()
        if device is None:
            return jsonify({'error': '设备不存在'}), 404
        
        conn.execute('DELETE FROM devices WHERE id = ?', (device_id,))
        conn.commit()
        return jsonify({'message': '设备删除成功', 'id': device_id})
    except Exception as e:
        return jsonify({'error': f'删除失败：{str(e)}'}), 500
    finally:
        conn.close()


@devices_bp.route('/api/import', methods=['POST'])
@login_required
def import_devices():
    """从Excel导入设备数据"""
    # 导入数据库连接函数
    from app import get_db_connection
    conn = get_db_connection()
    try:
        if 'file' not in request.files:
            return jsonify({'error': '没有文件上传'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '没有选择文件'}), 400
        
        if not file.filename.endswith('.xlsx'):
            return jsonify({'error': '只能上传Excel文件(.xlsx)'}), 400
        
        # 读取Excel文件
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        
        # 检查表头
        headers = []
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col).value
            headers.append(header)
        
        expected_headers = ['产品', '部门', '组别', '虚拟机名称', 'IP地址', '账号', '密码']
        # 允许第八列为备注字段
        if headers[:7] != expected_headers:
            return jsonify({'error': 'Excel文件格式错误，表头应为：产品、部门、组别、虚拟机名称、IP地址、账号、密码'}), 400
        
        success_count = 0
        error_count = 0
        errors = []
        
        # 导入数据
        for row in range(2, sheet.max_row + 1):
            product = sheet.cell(row=row, column=1).value
            department = sheet.cell(row=row, column=2).value
            group = sheet.cell(row=row, column=3).value
            vm_name = sheet.cell(row=row, column=4).value
            ip_address = sheet.cell(row=row, column=5).value
            username = sheet.cell(row=row, column=6).value
            password = sheet.cell(row=row, column=7).value
            
            # 读取备注字段（如果存在）
            remark = sheet.cell(row=row, column=8).value if sheet.max_column >= 8 else None
            
            # 检查是否是空行（所有必填字段都为空）
            if not vm_name and not ip_address and not username and not password:
                # 空行，跳过处理
                continue
            
            # 详细检查每个必填字段
            missing_fields = []
            if not vm_name:
                missing_fields.append('虚拟机名称')
            if not ip_address:
                missing_fields.append('IP地址')
            if not username:
                missing_fields.append('账号')
            if not password:
                missing_fields.append('密码')
            
            if missing_fields:
                error_count += 1
                errors.append(f'第{row}行数据不完整：缺少{"、".join(missing_fields)}')
                continue
            
            try:
                conn.execute(
                    'INSERT INTO devices (product, department, device_group, vm_name, ip_address, username, password, remark) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
                    (product, department, group, vm_name, ip_address, username, password, remark)
                )
                success_count += 1
            except Exception as e:
                error_count += 1
                if 'ip_address' in str(e):
                    errors.append(f'第{row}行IP地址({ip_address})已存在')
                else:
                    errors.append(f'第{row}行数据处理错误：{str(e)}')
        
        conn.commit()
        return jsonify({
            'message': '数据导入完成',
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors
        })
    except Exception as e:
        return jsonify({'error': f'导入失败：{str(e)}'}), 500
    finally:
        conn.close()


@devices_bp.route('/api/reset', methods=['POST'])
@admin_required
def reset_database():
    """重置数据库"""
    # 导入数据库连接函数
    from app import get_db_connection
    conn = get_db_connection()
    try:
        # 清空devices表
        conn.execute('DELETE FROM devices')
        
        # 重置自增ID
        conn.execute('DELETE FROM sqlite_sequence WHERE name = "devices"')
        
        conn.commit()
        return jsonify({'message': '数据库重置成功'})
    except Exception as e:
        return jsonify({'error': f'重置失败：{str(e)}'}), 500
    finally:
        conn.close()


@devices_bp.route('/api/vms', methods=['GET'])
@login_required
def get_all_vms():
    """获取所有虚拟机列表"""
    # 导入数据库连接函数
    from app import get_db_connection
    conn = get_db_connection()
    try:
        vms = conn.execute('SELECT id, ip_address, username, password FROM devices').fetchall()
        return jsonify([dict(vm) for vm in vms])
    finally:
        conn.close()
