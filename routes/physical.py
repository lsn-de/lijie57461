from flask import Blueprint, request, jsonify
from utils.decorators import login_required, admin_required
import openpyxl
import logging

# 创建物理机管理蓝图
physical_bp = Blueprint('physical', __name__)


@physical_bp.route('/api/physical-devices', methods=['GET'])
@login_required
def get_physical_devices():
    """获取所有物理机"""
    # 导入数据库连接函数
    from app import get_db_connection
    conn = get_db_connection()
    try:
        logging.debug('开始处理获取物理机请求')
        
        # 获取查询参数
        product = request.args.get('product')
        department = request.args.get('department')
        device_code = request.args.get('device_code')
        ip_address = request.args.get('ip_address')
        oob_ip = request.args.get('oob_ip')
        account = request.args.get('account')
        password = request.args.get('password')
        remark = request.args.get('remark')
        
        logging.debug(f'查询参数: product={product}, department={department}, device_code={device_code}, ip_address={ip_address}, oob_ip={oob_ip}, account={account}, password={password}')
        
        # 构建查询条件
        conditions = []
        params = []
        
        if product:
            conditions.append("product LIKE ?")
            params.append(f"%{product}%")
        if department:
            conditions.append("department LIKE ?")
            params.append(f"%{department}%")
        if device_code:
            conditions.append("device_code LIKE ?")
            params.append(f"%{device_code}%")
        if ip_address:
            conditions.append("ip_address LIKE ?")
            params.append(f"%{ip_address}%")
        if oob_ip:
            conditions.append("oob_ip LIKE ?")
            params.append(f"%{oob_ip}%")
        if account:
            conditions.append("account LIKE ?")
            params.append(f"%{account}%")
        if password:
            conditions.append("password LIKE ?")
            params.append(f"%{password}%")
        if remark:
            conditions.append("remark LIKE ?")
            params.append(f"%{remark}%")
        
        # 构建SQL查询
        sql = "SELECT * FROM physical_devices"
        if conditions:
            sql += " WHERE " + " AND ".join(conditions)
        sql += " ORDER BY id"
        
        logging.debug(f'执行SQL查询: {sql}, 参数: {params}')
        
        # 执行查询
        devices = conn.execute(sql, params).fetchall()
        logging.debug(f'查询结果数量: {len(devices)}')
        
        # 获取当前登录用户的信息
        from flask import session
        user_id = session.get('user_id')
        user = conn.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
        user_role = session.get('role')
        user_department = user['department'] if user else None
        
        result = []
        for device in devices:
            try:
                device_dict = {
                    'id': device['id'],
                    'product': device['product'],
                    'department': device['department'],
                    'device_code': device['device_code'],
                    'ip_address': device['ip_address'],
                    'oob_ip': device['oob_ip'],
                    'account': device['account'],
                    'remark': device['remark'],
                    'created_at': device['created_at']
                }
                
                # 只有管理员或同部门用户才能看到密码
                if user_role == 'admin' or (user_department and device['department'] == user_department):
                    device_dict['password'] = device['password']
                
                result.append(device_dict)
                logging.debug(f'处理设备数据: {device_dict}')
            except Exception as e:
                logging.error(f'处理设备数据时出错: {str(e)}')
                # 打印设备数据，看看具体是什么
                logging.error(f'设备数据: {device}')
                # 跳过这个设备，继续处理下一个
                continue
        
        logging.debug('返回物理机数据成功')
        return jsonify(result)
    except Exception as e:
        logging.error(f'获取物理机数据失败: {str(e)}')
        import traceback
        logging.error(traceback.format_exc())
        return jsonify({'error': f'获取物理机数据失败: {str(e)}'}), 500
    finally:
        conn.close()


@physical_bp.route('/api/physical-devices/<int:device_id>', methods=['GET'])
@login_required
def get_physical_device(device_id):
    """获取单个物理机"""
    # 导入数据库连接函数
    from app import get_db_connection
    conn = get_db_connection()
    try:
        device = conn.execute('SELECT * FROM physical_devices WHERE id = ?', (device_id,)).fetchone()
        
        # 获取当前登录用户的信息
        from flask import session
        user_id = session.get('user_id')
        user = conn.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
        user_role = session.get('role')
        user_department = user['department'] if user else None
        
        if device is None:
            return jsonify({'error': '物理机不存在'}), 404
        
        device_dict = {
                    'id': device['id'],
                    'product': device['product'],
                    'department': device['department'],
                    'device_code': device['device_code'],
                    'ip_address': device['ip_address'],
                    'oob_ip': device['oob_ip'],
                    'account': device['account'],
                    'remark': device['remark'],
                    'created_at': device['created_at']
                }
        
        # 只有管理员或同部门用户才能看到密码
        if user_role == 'admin' or (user_department and device['department'] == user_department):
            device_dict['password'] = device['password']
        
        return jsonify(device_dict)
    finally:
        conn.close()


@physical_bp.route('/api/physical-devices', methods=['POST'])
@login_required
def add_physical_device():
    """添加物理机"""
    # 导入数据库连接函数
    from app import get_db_connection
    conn = get_db_connection()
    try:
        # 支持 JSON 和表单数据
        if request.is_json:
            data = request.get_json()
        else:
            data = request.form
        
        # 检查必填字段
        if not all(key in data for key in ['ip_address', 'account', 'password', 'department']):
            return jsonify({'error': '缺少必要字段'}), 400

        # 获取其他字段，如果不存在则设为None
        product = data.get('product') or None
        department = data.get('department') or None
        device_code = data.get('device_code') or None
        oob_ip = data.get('oob_ip') or None
        remark = data.get('remark') or None

        # 插入数据
        cursor = conn.execute('''
            INSERT INTO physical_devices (product, department, device_code, ip_address, oob_ip, account, password, remark)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            product, department, device_code, data['ip_address'], oob_ip, data['account'], data['password'], remark
        ))
        
        # 获取新插入的ID
        new_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        conn.commit()
        return jsonify({'message': '物理机添加成功', 'id': new_id, 'ip_address': data['ip_address']})
    except Exception as e:
        return jsonify({'error': f'添加失败：{str(e)}'}), 500
    finally:
        conn.close()


@physical_bp.route('/api/physical-devices/<int:device_id>', methods=['PUT'])
@login_required
def update_physical_device(device_id):
    """更新物理机"""
    # 导入数据库连接函数
    from app import get_db_connection
    conn = get_db_connection()
    try:
        # 支持 JSON 和表单数据
        if request.is_json:
            data = request.get_json()
        else:
            data = request.form
        
        # 检查设备是否存在
        device = conn.execute('SELECT * FROM physical_devices WHERE id = ?', (device_id,)).fetchone()
        if device is None:
            return jsonify({'error': '物理机不存在'}), 404
        
        # 构建更新语句，只更新提供的字段
        update_fields = []
        update_params = []
        
        if 'product' in data:
            update_fields.append('product = ?')
            update_params.append(data['product'])
        if 'department' in data:
            update_fields.append('department = ?')
            update_params.append(data['department'])
        if 'device_code' in data:
            update_fields.append('device_code = ?')
            update_params.append(data['device_code'])
        if 'ip_address' in data:
            update_fields.append('ip_address = ?')
            update_params.append(data['ip_address'])
        if 'oob_ip' in data:
            update_fields.append('oob_ip = ?')
            update_params.append(data['oob_ip'])
        if 'account' in data:
            update_fields.append('account = ?')
            update_params.append(data['account'])
        if 'password' in data:
            update_fields.append('password = ?')
            update_params.append(data['password'])
        if 'remark' in data:
            update_fields.append('remark = ?')
            update_params.append(data['remark'])
        
        # 如果没有字段需要更新，直接返回成功
        if not update_fields:
            return jsonify({'message': '物理机更新成功'})
        
        # 构建 SQL 语句
        sql = f"UPDATE physical_devices SET {', '.join(update_fields)} WHERE id = ?"
        update_params.append(device_id)
        
        conn.execute(sql, update_params)
        conn.commit()
        return jsonify({'message': '物理机更新成功'})
    except Exception as e:
        return jsonify({'error': f'更新失败：{str(e)}'}), 500
    finally:
        conn.close()


@physical_bp.route('/api/physical-devices/<int:device_id>', methods=['DELETE'])
@admin_required
def delete_physical_device(device_id):
    """删除物理机"""
    # 导入数据库连接函数
    from app import get_db_connection
    conn = get_db_connection()
    try:
        # 检查设备是否存在
        device = conn.execute('SELECT * FROM physical_devices WHERE id = ?', (device_id,)).fetchone()
        if device is None:
            return jsonify({'error': '物理机不存在'}), 404
        
        conn.execute('DELETE FROM physical_devices WHERE id = ?', (device_id,))
        conn.commit()
        return jsonify({'message': '物理机删除成功', 'id': device_id})
    except Exception as e:
        return jsonify({'error': f'删除失败：{str(e)}'}), 500
    finally:
        conn.close()


@physical_bp.route('/api/physical-devices/import', methods=['POST'])
@login_required
def import_physical_devices():
    """导入物理机"""
    # 导入数据库连接函数
    from app import get_db_connection
    conn = get_db_connection()
    try:
        logging.debug('开始处理物理机导入请求')
        # 检查是否有文件上传
        if 'file' not in request.files:
            logging.error('没有文件上传')
            return jsonify({'error': '没有文件上传'}), 400
        
        file = request.files['file']
        logging.debug(f'获取到文件：{file.filename}')
        
        # 检查文件名是否为空
        if file.filename == '':
            logging.error('没有选择文件')
            return jsonify({'error': '没有选择文件'}), 400
        
        # 检查文件类型
        if not file.filename.endswith('.xlsx'):
            logging.error(f'文件类型错误：{file.filename}，只支持xlsx格式')
            return jsonify({'error': '只支持xlsx格式的文件'}), 400
        
        # 读取Excel文件
        logging.debug('开始读取Excel文件')
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        logging.debug(f'成功加载Excel文件，工作表：{sheet.title}')
        
        # 计算最大列数，确保行的完整性
        max_columns = sheet.max_column
        logging.debug(f'Excel文件最大列数：{max_columns}')
        
        if max_columns < 8:
            logging.error(f'Excel文件格式错误：至少需要8列数据，但只找到{max_columns}列')
            return jsonify({'error': f'Excel文件格式错误：至少需要8列数据，但只找到{max_columns}列'}), 400
        
        # 检查表头
        headers = []
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col).value
            headers.append(header)
        
        expected_headers = ['产品', '部门', '设备编码', 'IP地址', '带外IP', '账号', '密码', '备注']
        if headers[:8] != expected_headers:
            logging.error(f'Excel文件格式错误，表头应为：产品、部门、设备编码、IP地址、带外IP、账号、密码、备注')
            return jsonify({'error': 'Excel文件格式错误，表头应为：产品、部门、设备编码、IP地址、带外IP、账号、密码、备注'}), 400
        
        # 跳过表头行（第一行）
        success_count = 0
        error_count = 0
        errors = []
        
        logging.debug('开始处理Excel数据行')
        for row in range(2, sheet.max_row + 1):
            logging.debug(f'处理第{row}行数据')
            # 详细检查每个必填字段
            # 注意：根据Excel模板的实际格式，数据顺序为：产品, 部门, 设备编码, IP地址, 带外IP, 账号, 密码, 备注
            # 第一列：产品（例如：测试虚拟化资源池）
            # 第二列：部门（例如：投资服务PDU）
            # 第三列：设备编码
            # 第四列：IP地址
            # 第五列：带外IP
            # 第六列：账号
            # 第七列：密码
            # 第八列：备注
            product = sheet.cell(row=row, column=1).value
            department = sheet.cell(row=row, column=2).value
            device_code = sheet.cell(row=row, column=3).value
            ip_address = sheet.cell(row=row, column=4).value
            oob_ip = sheet.cell(row=row, column=5).value
            account = sheet.cell(row=row, column=6).value
            password = sheet.cell(row=row, column=7).value
            remark = sheet.cell(row=row, column=8).value if sheet.max_column >= 8 else None
            
            # 处理空值
            product = product if product is not None else '未分类'
            device_code = device_code if device_code is not None else 'N/A'
            oob_ip = oob_ip if oob_ip is not None else 'N/A'
            remark = remark if remark is not None else 'N/A'
            account = account if account is not None else ''
            department = department if department is not None else ''
            password = password if password is not None else ''
            
            logging.debug(f'处理后的数据：product={product}, department={department}, device_code={device_code}, ip_address={ip_address}, oob_ip={oob_ip}, account={account}, password={password}, remark={remark}')
            
            # 检查是否是空行（所有必填字段都为空）
            if not ip_address and not account and not password and not department:
                # 空行，跳过处理
                continue
            
            # 检查必填字段（只有IP地址是必填的，其他字段都可以为空）
            missing_fields = []
            if not ip_address:
                missing_fields.append('IP地址')
            
            if missing_fields:
                logging.debug(f'第{row}行：必填字段不完整，跳过')
                error_count += 1
                errors.append(f'第{row}行数据不完整：缺少{"、".join(missing_fields)}')
                continue
            
            # 插入数据到数据库
            logging.debug(f'第{row}行：准备插入数据库')
            # 注意：根据Excel模板的实际格式，数据顺序为：产品, 部门, 设备编码, IP地址, 带外IP, 账号, 密码, 备注
            logging.debug(f'插入数据：product={product}, department={department}, device_code={device_code}, ip_address={ip_address}, oob_ip={oob_ip}, account={account}, password={password}, remark={remark}')
            try:
                conn.execute('''
                    INSERT INTO physical_devices (product, department, device_code, ip_address, oob_ip, account, password, remark)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    product,  # product
                    department,  # department
                    device_code,  # device_code
                    ip_address,  # ip_address
                    oob_ip,  # oob_ip
                    account,  # account
                    password,  # password
                    remark  # remark
                ))
                
                success_count += 1
                logging.debug(f'第{row}行：数据插入成功')
            except Exception as e:
                error_msg = str(e)
                if 'ip_address' in error_msg:
                    errors.append(f'第{row}行IP地址({ip_address})已存在')
                elif 'NOT NULL constraint failed' in error_msg:
                    if 'physical_devices.account' in error_msg:
                        errors.append(f'第{row}行账号不能为空')
                    elif 'physical_devices.department' in error_msg:
                        errors.append(f'第{row}行部门不能为空')
                    elif 'physical_devices.password' in error_msg:
                        errors.append(f'第{row}行密码不能为空')
                    else:
                        errors.append(f'第{row}行数据不完整：缺少必填字段')
                else:
                    errors.append(f'第{row}行数据处理错误：{error_msg}')
                error_count += 1
        
        conn.commit()
        logging.debug(f'数据行处理完成，成功：{success_count}，错误：{error_count}')
        return jsonify({
            'message': '物理机导入完成',
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors
        })
    except Exception as e:
        logging.error(f'导入失败：{str(e)}')
        import traceback
        logging.error(traceback.format_exc())
        return jsonify({'error': f'导入失败：{str(e)}'}), 500
    finally:
        conn.close()


@physical_bp.route('/api/physical-reset', methods=['POST'])
@admin_required
def reset_physical_database():
    """重置物理机数据库"""
    # 导入数据库连接函数
    from app import get_db_connection
    conn = get_db_connection()
    try:
        # 清空physical_devices表
        conn.execute('DELETE FROM physical_devices')
        
        # 重置自增ID
        conn.execute('DELETE FROM sqlite_sequence WHERE name = "physical_devices"')
        
        conn.commit()
        return jsonify({'message': '物理机数据库重置成功'})
    except Exception as e:
        return jsonify({'error': f'重置失败：{str(e)}'}), 500
    finally:
        conn.close()


@physical_bp.route('/api/physical-devices-all', methods=['GET'])
@login_required
def get_all_physical_devices():
    """获取所有物理机列表"""
    # 导入数据库连接函数
    from app import get_db_connection
    conn = get_db_connection()
    try:
        devices = conn.execute('SELECT id, ip_address, account, password FROM physical_devices').fetchall()
        return jsonify([dict(device) for device in devices])
    finally:
        conn.close()


@physical_bp.route('/api/physical-devices-for-monitor', methods=['GET'])
@login_required
def get_physical_devices_for_monitor():
    """获取所有物理机列表（用于监控）"""
    # 导入数据库连接函数
    from app import get_db_connection
    conn = get_db_connection()
    try:
        devices = conn.execute('SELECT id, ip_address, account, password FROM physical_devices').fetchall()
        return jsonify([dict(device) for device in devices])
    finally:
        conn.close()
